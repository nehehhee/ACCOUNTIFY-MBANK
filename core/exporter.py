"""
Zápis do Excelu — Transakcie (Excel tabuľka) + Protiúčty (analýza).
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Farby ──────────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1a1a2e")
HDR_FONT   = Font(bold=True, color="e2e8f0", size=10)
GREEN_FILL = PatternFill("solid", fgColor="1a2e1a")
GREEN_FONT = Font(bold=True, color="4ade80", size=10)
THIN       = Side(style="thin", color="3a3a3a")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── Šírky stĺpcov pre každý hárok ─────────────────────────────────────────────
TRANSAKCIE_WIDTHS = {
    "banka":           10,
    "nazov_uctu":      22,
    "iban_uctu":       28,
    "datum":           13,
    "suma":            12,
    "cd":               5,
    "nazov_protiuctu": 28,
    "iban_protiuctu":  28,
    "typ":             22,
    "poznamka":        35,
}

PROTIUCTY_WIDTHS = {
    "IBAN protiúčtu":      28,
    "Názov protiúčtu":     28,
    "Počet transakcií":    18,
    "Odoslané (D)":        16,
    "Prijaté (C)":         16,
    "Bilancia":            16,
    "Posledná transakcia": 20,
}


def write_excel(df: pd.DataFrame, path: str):
    """Zapíše df do xlsx s dvoma hárkami: Transakcie a Protiúčty."""
    wb = Workbook()
    wb.remove(wb.active)

    _write_transakcie(wb, df)
    _write_protiucty(wb, df)

    try:
        wb.save(path)
    except OSError as e:
        raise RuntimeError(f"Nepodarilo sa uložiť Excel:\n{e}") from e


# ── Hárok 1: Transakcie ────────────────────────────────────────────────────────

def _write_transakcie(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Transakcie")

    # Dáta
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            if r_idx == 1:
                cell.font = HDR_FONT
                cell.fill = HDR_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Excel tabuľka (umožní kontingenčnú tabuľku jedným klikom)
    last_col = get_column_letter(len(df.columns))
    last_row = len(df) + 1
    tbl = Table(
        displayName="Transakcie",
        ref=f"A1:{last_col}{last_row}",
    )
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)

    # Šírky stĺpcov
    for c_idx, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = TRANSAKCIE_WIDTHS.get(col_name, 16)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


# ── Hárok 2: Protiúčty ────────────────────────────────────────────────────────

def _write_protiucty(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Protiúčty")

    summary = _build_protiucty(df)

    headers = list(PROTIUCTY_WIDTHS.keys())

    # Hlavička
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.font = GREEN_FONT
        cell.fill = GREEN_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Dáta
    for r_idx, row in enumerate(summary.itertuples(index=False), start=2):
        values = list(row)
        for c_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")
            # Červená/zelená pre D/C sumy
            if c_idx == 4 and value:
                cell.font = Font(color="f87171")
            if c_idx == 5 and value:
                cell.font = Font(color="4ade80")

    # Excel tabuľka
    last_col = get_column_letter(len(headers))
    last_row = len(summary) + 1
    if last_row > 1:
        tbl = Table(
            displayName="Protiucty",
            ref=f"A1:{last_col}{last_row}",
        )
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium6",
            showRowStripes=True,
        )
        ws.add_table(tbl)

    # Šírky
    for c_idx, col_name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = PROTIUCTY_WIDTHS[col_name]

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def _build_protiucty(df: pd.DataFrame) -> pd.DataFrame:
    """Agreguje transakcie podľa protiúčtu."""
    if df.empty:
        return pd.DataFrame(columns=list(PROTIUCTY_WIDTHS.keys()))

    # Pracujeme len s riadkami kde je IBAN protiúčtu
    sub = df[df["iban_protiuctu"].astype(str).str.strip() != ""].copy()
    if sub.empty:
        return pd.DataFrame(columns=list(PROTIUCTY_WIDTHS.keys()))

    sub["suma"] = pd.to_numeric(sub["suma"], errors="coerce").fillna(0)

    grouped = sub.groupby(["iban_protiuctu", "nazov_protiuctu"], dropna=False)

    pocet     = grouped["suma"].count().rename("pocet")
    odoslane  = sub[sub["cd"] == "D"].groupby(
        ["iban_protiuctu", "nazov_protiuctu"])["suma"].sum().rename("odoslane")
    prijate   = sub[sub["cd"] == "C"].groupby(
        ["iban_protiuctu", "nazov_protiuctu"])["suma"].sum().rename("prijate")
    posledna  = grouped["datum"].max().rename("posledna")

    result = pd.concat([pocet, odoslane, prijate, posledna], axis=1).reset_index()
    result["odoslane"] = result["odoslane"].fillna(0).round(2)
    result["prijate"]  = result["prijate"].fillna(0).round(2)
    result["bilancia"] = (result["prijate"] - result["odoslane"]).round(2)
    result = result.sort_values("pocet", ascending=False)

    result.columns = list(PROTIUCTY_WIDTHS.keys())
    return result
